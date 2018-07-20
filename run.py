import pandas as pd
from pandas import ExcelWriter
import openpyxl
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors
from openpyxl.cell import Cell
import numpy as np
import math
import requests
import json
import datetime
import re
from bs4 import BeautifulSoup
import time
import warnings
import plotly.plotly as py
warnings.filterwarnings("ignore")

# NOAA token request page: https://www.ncdc.noaa.gov/cdo-web/token
noaa_api_token = ''


MAX_RETRIES = 20

session = requests.Session()
adapter = requests.adapters.HTTPAdapter(max_retries=MAX_RETRIES, pool_connections=10)
session.mount('https://', adapter)
session.mount('http://', adapter)
session.headers.update({'token': noaa_api_token})

states_abb = ['CT','DE','ME','MD','MA','NH','NJ','NY','PA','RI','VT','IA','MI','MN','WI','IL','IN','KY','MO','OH','TN','WV','AL','FL','GA','NC','SC','VA','MT','NE','ND','SD','WY','AR','KS','LA','MS','OK','TX','AZ','CO','NM','UT','ID','OR','WA','CA','NV']
state_names = ['Connecticut','Delaware','Maine','Maryland','Massachusetts','New Hampshire','New Jersey','New York','Pennsylvania','Rhode Island','Vermont','Iowa','Michigan','Minnesota','Wisconsin','Illinois','Indiana','Kentucky','Missouri','Ohio','Tennessee','West Virginia','Alabama','Florida','Georgia','North Carolina','South Carolina','Virginia','Montana','Nebraska','North Dakota','South Dakota','Wyoming','Arkansas','Kansas','Louisiana','Mississippi','Oklahoma','Texas','Arizona','Colorado','New Mexico','Utah','Idaho','Oregon','Washington','California','Nevada']

excel_file_name = 'FIPSTemperatureFeed5.0.xlsx'

united_states_population = 321198379

GFill = PatternFill(start_color='dbd9d9',
                   end_color='dbd9d9',
                   fill_type='solid')

LGFill = PatternFill(start_color='f0f0f0',
                   end_color='f0f0f0',
                   fill_type='solid')

LightOrangeFill = PatternFill(start_color='ffe4b3',
                   end_color='ffe4b3',
                   fill_type='solid')

bottom_border = Border(bottom=Side(style='thin'))

right_border = Border(right=Side(style='thin'))

bottom_right_border = Border(bottom=Side(style='thin'),
                            right=Side(style='thin'))

bold_font = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

output_font = Font(name='Calibri',
                size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='ed8b13')

center_align = Alignment(horizontal='center')


def get_response(url):
    
    response = session.get(url)
    response.raise_for_status()
    
    return response
    
def check_add_date(date):
    
    '''
    Will check column 'A' for the date that is passed in.
    It will pass if the date is already in the column and add the date if it is not already in the columns
    '''
    
    FIPSTemperatureFeed = openpyxl.load_workbook(excel_file_name)
    
    for sheet_name in FIPSTemperatureFeed.sheetnames:
        
        if sheet_name not in ['Recent Map']:

            temperature_sheet = FIPSTemperatureFeed[sheet_name]

            date_column = temperature_sheet['A']

            dates_in_date_column = [date_cell.value for date_cell in date_column if date_cell.value is not None]

            if date in dates_in_date_column:

                pass

            else:

                next_row = len(dates_in_date_column) + 1
                indexing_cell = 'A'+str(next_row)

                cell_for_date = temperature_sheet[indexing_cell]

                # Adding the date to column 'A'
                cell_for_date.value = date

                cell_for_date.border = right_border
                cell_for_date.fill = LGFill
                cell_for_date.alignment = center_align
        else:
            
            pass

    FIPSTemperatureFeed.save(excel_file_name)
    
def input_cell(date_input, header_value, cell_input):
    
    '''
    date_input: Date to locate row for
    header_value: Value in row 1 of 'State Averages' sheet that you want to add a value for the date_input
    cell_input: Input value to add to the cell where the date_input column and header_value row intersect
    '''
    
    FIPSTemperatureFeed = openpyxl.load_workbook(excel_file_name)
    sheet = FIPSTemperatureFeed['State Averages']
    
    for cell in sheet[1]:

        if cell.value == 'Date':

            date_col = cell.column

        else:

            pass

    for cell in sheet[date_col]:

        if cell.value == date_input:

            date_row = cell.row

        else:

            pass
    
    for cell in sheet[1]:
        
        if cell.value == header_value:
            
            header_col = cell.col_idx
            
    in_cell = sheet.cell(row=date_row, column=header_col)
    in_cell.value = cell_input
    in_cell.alignment = center_align
    in_cell.number_format = '0.000'
    
    if header_value in ['Population Weighted Minimum','Population Weighted Average','Population Weighted Maximum']:
        
        in_cell.font = output_font
        
    else:
        
        pass
    
    FIPSTemperatureFeed.save(excel_file_name)
    
def get_state_name(abbreviation):
    
    '''
    abbreviation: Two letter state abbreviation to get state name for
    '''
    
    for num, abb in enumerate(states_abb):
        
        if abb.upper() == abbreviation.upper():
        
            return state_names[num]
    
        else:
            
            pass
        
def get_state_abb(state_name):
    
    '''
    abbreviation: Two letter state abbreviation to get state name for
    '''
    
    for num, state in enumerate(state_names):
        
        if state_name == state:
            
            return states_abb[num]
            
        else:
            
            pass
        
recent_date_response = get_response('https://www.ncdc.noaa.gov/cdo-web/api/v2/datasets?datatypeid=TOBS')
recent_date_id = recent_date_response.json()['results'][0]['id']

if recent_date_id == 'GHCND':
    
    recent_date = recent_date_response.json()['results'][0]['maxdate']
    recent_date_datetime = datetime.datetime(int(recent_date.split('-')[0]), int(recent_date.split('-')[1]), int(recent_date.split('-')[2]))
    
else:
    
    print('Recent date unavailable')
    
check_add_date(recent_date_datetime)

population_dataframe = pd.read_excel('PopulationByFIPS.xlsx')
population_dataframe['Fips'] = [str(fips_code).zfill(5) for fips_code in population_dataframe['Fips']]

response = get_response('https://www.ncdc.noaa.gov/cdo-web/api/v2/locations?datasetid=GHCND&limit=1000')

response_result_set = response.json()['metadata']['resultset']

amount_iters = math.ceil(response_result_set['count'] / response_result_set['limit'])

code_name_tups = []

for num in range(amount_iters):

    response = get_response('https://www.ncdc.noaa.gov/cdo-web/api/v2/locations?datasetid=GHCND&limit=1000&offset='+str(num)+'000')
    
    # Will pass if there are no results
    try:

        # Gets all ghcnds that have 100% data coverage, and are in the United States
        [code_name_tups.append((tup[0].split('FIPS:')[1],tup[1])) for tup in [(dic['id'],dic['name']) for dic in response.json()['results'] if dic['datacoverage'] == 1 and dic['name'][-2:] in states_abb] if tup[0].startswith('FIPS') and tup[0].split('FIPS:')[1] in list(population_dataframe['Fips'])]

    except KeyError:
        
        pass
    
fips_list, county_names_list, state_names_list = [tup[0] for tup in code_name_tups], [tup[1].split(',')[0] for tup in code_name_tups], [tup[1].split(',')[1][1:] for tup in code_name_tups]

class County(object):
    
    def __init__(self, fips_code, county_name, state_name, county_population):
        
        self.fips_code = fips_code
        self.county_name = county_name
        self.state_name = state_name
        self.county_population = county_population
        
        # Calculating weight based off of percent from whole us population
        self.us_weight = self.county_population / united_states_population
        
        
        # Pulling data from the API
        response = get_response('https://www.ncdc.noaa.gov/cdo-web/api/v2/data?datasetid=GHCND&locationid=FIPS:'+self.fips_code+'&startdate='+str(recent_date)+'&enddate='+str(recent_date)+'&units=standard&limit=1000')
        
        try:

            json = pd.DataFrame(response.json()['results'])
            json = json[json['datatype'] == 'TOBS']

            try:
                
                fips_min_temp_for_day = float(np.min([x for x in json['value'] if not math.isnan(x)]))
                
            except ValueError:
                
                fips_min_temp_for_day = np.nan
                
            try:
                
                fips_average_temp_for_day = float(np.average([x for x in json['value'] if not math.isnan(x)]))
                
            except ValueError:
                
                fips_average_temp_for_day = np.nan
                
            try:
                
                fips_max_temp_for_day = float(np.max([x for x in json['value'] if not math.isnan(x)]))
            
            except ValueError:
                
                fips_max_temp_for_day = np.nan

            
            self.min_county_temperature = fips_min_temp_for_day
            self.county_temperature = fips_average_temp_for_day
            self.max_county_temperature = fips_max_temp_for_day

        except:

            self.min_county_temperature = np.nan
            self.county_temperature = np.nan
            self.max_county_temperature = np.nan
            
county_classes = [County(fips, county, state, list(population_dataframe[population_dataframe['Fips'] == fips]['Population'])[0]) for fips, county, state in zip(fips_list, county_names_list, state_names_list)]

state_weight_average = {}

for state in states_abb:
    
    state_population, state_min_temperatures, state_avg_temperatures, state_max_temperatures = 0, [], [], []
    
    for county in county_classes:
        
        if county.state_name == state:
            
            state_population += county.county_population
            
            state_min_temperatures.append(county.min_county_temperature)
            state_avg_temperatures.append(county.county_temperature)
            state_max_temperatures.append(county.max_county_temperature)
            
        else:
            
            pass
        
        
        try:
            
            state_minimum = np.min([temperature for temperature in state_min_temperatures if not math.isnan(temperature)])
            
        except ValueError:
            
            state_minimum = np.nan
        
        try:
            
            state_average = np.average([temperature for temperature in state_avg_temperatures if not math.isnan(temperature)])
            
        except ValueError:
            
            state_average = np.nan
            
        try:
            
            state_maximum = np.max([temperature for temperature in state_max_temperatures if not math.isnan(temperature)])
            
        except:
            
            state_maximum = np.nan
        
    # The total state population added earlier when iterating through counties in county_classes divided by the united states population as a whole
    state_weight = state_population / united_states_population
            
    # Adding to the 'state_weight_average' dict. Will be a tuple of (state weight, [min, avg, max])
    state_weight_average[state] = (state_weight, [state_minimum, state_average, state_maximum])
    
weighted_mins, weighted_avgs, weighted_maxes = [], [], []

for state, weight_temperature_tuple in state_weight_average.items():
    
    weight, temperature_list = weight_temperature_tuple[0], weight_temperature_tuple[1]
    
    weighted_temperatures = [temperature*weight for temperature in temperature_list]
    
    weighted_mins.append(weighted_temperatures[0])
    weighted_avgs.append(weighted_temperatures[1])
    weighted_maxes.append(weighted_temperatures[2])
    
population_weighted_min, population_weighted_avg, population_weighted_max = np.sum([temperature for temperature in weighted_mins if not math.isnan(temperature)]), np.sum([temperature for temperature in weighted_avgs if not math.isnan(temperature)]), np.sum([temperature for temperature in weighted_maxes if not math.isnan(temperature)])

saved_population_weighted_vars = {
    'Population Weighted Minimum':population_weighted_min,
    'Population Weighted Average':population_weighted_avg,
    'Population Weighted Maximum':population_weighted_max,
}

for state_abbreviation, weight_temperature_tuple in state_weight_average.items():
    
    saved_population_weighted_vars[get_state_name(state_abbreviation) + ' Minimum'], saved_population_weighted_vars[get_state_name(state_abbreviation) + ' Average'], saved_population_weighted_vars[get_state_name(state_abbreviation) + ' Maximum'] = weight_temperature_tuple[1][0], weight_temperature_tuple[1][1], weight_temperature_tuple[1][2]
    
[input_cell(recent_date_datetime, string, var) for string, var in saved_population_weighted_vars.items()]

avg_temp_dict_by_state = {key.split(' Average')[0]:temperature for key, temperature in saved_population_weighted_vars.items() if 'Average' in key and key not in ['Population Weighted Minimum','Population Weighted Average','Population Weighted Maximum']}

# Creating dataframe to use for plotting statewide data
frame = pd.DataFrame(avg_temp_dict_by_state, index=['temperature']).T
frame.index = [get_state_abb(val) for val in frame.index]
frame['code'] = frame.index
frame.reset_index(inplace=True, drop=True)

for col in frame.columns:
    frame[col] = frame[col].astype(str)

scl = [[0.0, 'rgb(242,240,247)'],[0.2, 'rgb(218,218,235)'],[0.4, 'rgb(188,189,220)'],\
            [0.6, 'rgb(158,154,200)'],[0.8, 'rgb(117,107,177)'],[1.0, 'rgb(84,39,143)']]

data = [ dict(
        type='choropleth',
        colorscale = scl,
        autocolorscale = False,
        locations = frame['code'],
        z = frame['temperature'].astype(float),
        locationmode = 'USA-states',
        marker = dict(
            line = dict (
                color = 'rgb(255,255,255)',
                width = 2
            ) ),
        colorbar = dict(
            title = "Degrees (F)")
        ) ]

layout = dict(
        title = 'Temperature by State for {} <br>(Hover for breakdown)'.format(recent_date),
        geo = dict(
            scope='usa',
            projection=dict( type='albers usa' ),
            showlakes = True,
            lakecolor = 'rgb(255, 255, 255)'),
             )
    
fig = dict( data=data, layout=layout )
py.iplot( fig, filename='Temperature by State for {}'.format(recent_date))
